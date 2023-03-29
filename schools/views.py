from datetime import datetime, timedelta
import secrets
import string

from email import message
from django.urls import reverse
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count
from schedule.models import Assessment, Attendance, ProfTest, TrainingCycle, \
                            TrainingStream
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.utils.encoding import escape_uri_path

from users.models import User
from regions.models import City, TerAdministration, Address
from .models import Grade, School, SchoolContactPersone, SchoolStudent

from education_centers.forms import ImportDataForm
from users.mailing import send_mail
from . import imports

#ЛК школы
@csrf_exempt
@login_required
def school_profile(request, school_id):
    title = "Мой Выбор | Личный кабинет"
    school = get_object_or_404(School, id=school_id)
    message = ""
    if request.method == "POST":
        if 'exclude-student' in request.POST:
            student_id = request.POST[f'student_id']
            student = SchoolStudent.objects.get(id=student_id)
            test_id = request.POST[f'test_id']
            test = ProfTest.objects.get(id=test_id)
            test.students.remove(student)
            test.save()
        if 'edit-school' in request.POST:
            school_name = request.POST[f'school_name']
            inn = request.POST[f'inn']
            school.name = school_name
            school.inn = inn
            school_address = school.address
            city_id = request.POST[f'city']
            city = get_object_or_404(City, id=city_id)
            street = request.POST[f'street']
            building_number = request.POST[f'building_number']
            school_address.city = city
            school_address.street = street
            school_address.building_number = building_number
            user = school.school_contact.user
            last_name = request.POST[f'last_name']
            first_name = request.POST[f'first_name']
            middle_name = request.POST[f'middle_name']
            phone_number = request.POST[f'phone_number']
            email = request.POST[f'email']
            user.last_name = last_name.title()
            user.first_name = first_name.title()
            user.middle_name = middle_name.title()
            user.phone_number = phone_number.replace(" ", "")
            user.email = email
            user.save()
            school_address.save()
            school.save()
    contact = get_object_or_404(SchoolContactPersone, school=school.id)
    
    students = school.students.all()
    students_count = len(students)
    enroled_students_count = len(students.exclude(cycles=None))
    test_count = len(Attendance.objects.filter(student__in=students, is_attend=True))

    streams = TrainingStream.objects.filter(students__in=students)
    two_weeks = datetime.today() + timedelta(14)
    tests = ProfTest.objects.filter(stream__in=streams, date__gte=datetime.today(), date__lte=two_weeks).order_by('date', 'start_time')

    return render(request, "schools/school_profile.html",{
        "title": title,
        "cities": City.objects.all(),
        "school": school,
        "students_count": students_count,
        "enroled_students_count": enroled_students_count,
        "test_count": test_count,
        'tests': tests,
        'contact': contact,
        'date_today': datetime.today(),
        'date_two_weeks': two_weeks,
        "message": message
    })

@login_required
def grades_list(request, school_id):
    title = "Мой Выбор | Классы"
    school = get_object_or_404(School, id=school_id)
    contact = get_object_or_404(SchoolContactPersone, school=school.id)

    grades = Grade.objects.filter(school=school).order_by('grade', 'grade_letter')
    grades_list = {}
    for grade in range(6, 12):
        if len(grades.filter(grade=grade)) != 0:
            grades_list[grade] = grades.filter(grade=grade)

    return render(request, "schools/grades_list.html",{
        "title": title,
        "school": school,
        'contact': contact,
        "grades": grades_list,
    })

@csrf_exempt
@login_required
def grade(request, school_id, grade_id):
    title = "Мой Выбор | Класс"
    school = get_object_or_404(School, id=school_id)
    contact = get_object_or_404(SchoolContactPersone, school=school.id)
    
    grade = get_object_or_404(Grade, id=grade_id)
    message = ""
    if request.method == "POST":
        student_id = request.POST["student_id"]
        student = get_object_or_404(SchoolStudent, id=student_id)
        if 'cancel-stream' in request.POST:
            stream_id = request.POST["stream_id"]
            stream = get_object_or_404(TrainingStream, id=stream_id)

            stream.students.remove(student)
            stream.save()
            stream.cycle.students.remove(student)
            stream.cycle.save()
            for test in stream.tests.all():
                test.students.remove(student)
                test.save()
            assessment = Assessment.objects.filter(student=student, test__in=stream.tests.all())
            assessment.delete()
            attendance = Attendance.objects.filter(student=student, test__in=stream.tests.all())
            attendance.delete()
            message = "Success"
        if 'change-student' in request.POST:
            user = student.user
            first_name = request.POST["first_name"]
            middle_name = request.POST["middle_name"]
            last_name = request.POST["last_name"]
            email = request.POST["email"]
            phone = request.POST["phone"]
            user.first_name = first_name
            user.middle_name = middle_name
            user.last_name = last_name
            user.email = email
            user.phone = phone
            user.save()
            new_grade, new = Grade.objects.get_or_create(
                school=school,
                grade=int(request.POST["school_class"]),
                grade_letter = request.POST["school_class_latter"].capitalize()
            )
            student.grade = new_grade
            student.save()
            if len(grade.students.all()) == 0:
                grade.delete()
                grade = new_grade
            
            message = "Success"
    return render(request, "schools/grade.html",{
        "title": title,
        "school": school,
        'contact': contact,
        "grade": grade,
        "message": message
    })

@csrf_exempt
@login_required
def streams_enroll(request, school_id, grade_id):
    title = "Мой Выбор | Запись"
    school = get_object_or_404(School, id=school_id)
    contact = get_object_or_404(SchoolContactPersone, school=school.id)
    grade = get_object_or_404(Grade, id=grade_id)
    message = "-"
    if request.method == "POST":
        stream_id = request.POST['stream_id']
        stream = get_object_or_404(TrainingStream, id=stream_id)
        students = request.POST.getlist("students")
        stream.students.add(*students)
        stream.cycle.students.add(*students)
        stream.save()
        for student_id in students:
            student = get_object_or_404(SchoolStudent, id=student_id)
            for test in stream.tests.all():
                test.students.add(*stream.students.all())
                test.save()
                for criterion in test.program.criteria.all():
                    assessment = Assessment(
                        test=test,
                        student=student,
                        criterion=criterion
                    )
                    assessment.save()
                for criterion in test.program.soft_criteria.all():
                    assessment = Assessment(
                        test=test,
                        student=student,
                        criterion=criterion
                    )
                    assessment.save()
                attendance = Attendance(
                    test=test,
                    student=student
                )
                attendance.save()
        message = "Success"
    cycles = TrainingCycle.objects.filter(schools=school)
    streams = TrainingStream.objects.filter(cycle__in=cycles)
    tests = ProfTest.objects.filter(stream__in=streams).exclude(date=None)
    streams = TrainingStream.objects.filter(tests__in=tests)

    students = SchoolStudent.objects.filter(school=school)

    return render(request, "schools/streams_enroll.html",{
        "title": title,
        "school": school,
        'contact': contact,
        "grade": grade,
        "students": students,
        "cycles": cycles,
        "streams": streams,
        "message": message
    })

@csrf_exempt
@login_required
def school_tests_list(request, school_id):
    if request.method == "POST" and 'exclude-student' in request.POST:
        student_id = request.POST[f'student_id']
        student = SchoolStudent.objects.get(id=student_id)
        test_id = request.POST[f'test_id']
        test = ProfTest.objects.get(id=test_id)
        test.students.remove(student)
        test.save()
    title = "Мой Выбор | Расписание проб"
    school = get_object_or_404(School, id=school_id)
    contact = get_object_or_404(SchoolContactPersone, school=school.id)

    students = school.students.all()
    tests = ProfTest.objects.filter(students__in=students).exclude(date=None).order_by('-date', '-start_time').distinct()

    return render(request, "schools/tests_list.html",{
        "title": title,
        "school": school,
        'contact': contact,
        "tests": tests,
    })

@login_required
def streams_list(request, school_id):
    title = 'Мой Выбор | Циклы'
    school = get_object_or_404(School, id=school_id)
    contact = get_object_or_404(SchoolContactPersone, school=school.id)

    cycles = TrainingCycle.objects.filter(schools=school)

    return render(request, "schools/streams_list.html",{
        "title": title,
        "school": school,
        'contact': contact,
        "cycles": cycles,
    })

# Изменение данных школы/конт. лица
@login_required
@csrf_exempt
def change_school(request):
    if request.method == "POST":
        try:
            school = School.objects.get(id=request.POST["school_id"])
        except School.DoesNotExist:
            return HttpResponseRedirect(reverse("school_profile", args=(school.id,)))
        school.name = request.POST["name"]
        try:
            school.city = City.objects.get(id=request.POST["city"])
        except City.DoesNotExist:
            pass
        school.adress = request.POST["adress"]
        school.save()
        try:
            contact = SchoolContactPersone.objects.get(school=school)
        except SchoolContactPersone.DoesNotExist:
            return HttpResponseRedirect(reverse("school_profile", args=(school.id,)))
        contact.user.phone_number = request.POST["phone"]
        contact.user.email = request.POST["email"]
        contact.user.save()
    return HttpResponseRedirect(reverse("school_profile", args=(school.id,)))

def password_generator():
    alphabet = string.ascii_letters + string.digits
    password = ''.join(secrets.choice(alphabet) for i in range(12))
    return password

# Добавление школ
@csrf_exempt
@login_required
def add_school(request):
    message = None
    if request.method == 'POST':
        inn = request.POST["INN"]
        duplicates = School.objects.filter(inn=inn)
        if len(duplicates) != 0:
            message = "Школа с таким ИНН уже существует!"
        else:
            city_id = request.POST["City"]
            city = City.objects.get(id=city_id)
            street = request.POST["Street"]
            building_number = request.POST["BuildingNumber"]

            address = Address(
                city=city,
                street=street,
                building_number=building_number
            )
            address.save()
            ter_admin_id = request.POST["TerAdmin"]
            ter_admin = TerAdministration.objects.get(id=ter_admin_id)
            school_name = request.POST["SchoolName"]
            school = School(
                inn=inn,
                name=school_name,
                ter_administration=ter_admin,
                address=address
            )
            school.save()
            email = request.POST["Email"]
            password = password_generator()
            user = User.objects.create_user(email, password)
            user.first_name = request.POST["FirstName"]
            user.middle_name = request.POST["MiddleName"]
            user.last_name = request.POST["LastName"]
            user.phone_number = request.POST["Phone"]
            user.role = 'RSC'
            user.save()
            contact = SchoolContactPersone(
                user=user,
                school=school
            )
            contact.save()
            message = "Success"
            
            #Отправляем email+пароль на почту
            subject = 'Данные для входа в личный кабинет skillsguide.ru'
            html = f'Здравствуйте, {user.first_name}!<p>Вам предоставлен доступ к платформе http://skillsguide.ru/ (проект "Мой выбор"), как представителю школы "{school.name}".</p> <p><br><b>Логин:</b> {user.email}<br><b>Пароль:</b> {password}</p><br><br>Это автоматическое письмо на него не нужно отвечать.'
            text = f'Здравствуйте!\n Здравствуйте, {user.first_name}! \nВам предоставлен доступ к платформе http://skillsguide.ru/ (проект "Мой выбор"), как представителю школы "{school.name}".\nЛогин: {user.email}\nПароль: {password} \n\nЭто автоматическое письмо на него не нужно отвечать.'
            to_name = f"{user.first_name} {user.last_name}"
            to_email = email
            send_mail(subject, html, text, to_name, to_email)

    cities = City.objects.all()
    ter_admins = TerAdministration.objects.all()
    
    return render(request, 'schools/add_school.html', {
        'cities': cities,
        'ter_admins': ter_admins,
        'message': message
    })

#Импорт школ
@csrf_exempt
@login_required
def import_schools(request):
    if request.method == 'POST':
        form = ImportDataForm(request.POST, request.FILES)
        if form.is_valid():
            data = imports.schools(form)
            message = data[0]
            return render(request, 'schools/import_schools.html', {
                'form': ImportDataForm(),
                'message': message,
                'data': data
            })
        else:
            data = form.errors
            message = "IndexError"

    return render(request, 'schools/import_schools.html', {
        'form' : ImportDataForm()
    })

def export_students(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Участники проб"
    ws.cell(row=1, column=1, value="Имя")
    ws.cell(row=1, column=2, value="Отчество")
    ws.cell(row=1, column=3, value="Фамилия")
    ws.cell(row=1, column=4, value="Дата рождения")
    ws.cell(row=1, column=5, value="Email")
    ws.cell(row=1, column=6, value="Телефон")
    ws.cell(row=1, column=7, value="Школа")
    ws.cell(row=1, column=8, value="Класс")
    attendancies = Attendance.objects.filter(is_attend=True)
    students = SchoolStudent.objects.filter(
        attendance__in=attendancies).distinct()
    row = 2
    for student in students:
        ws.cell(row=row, column=1, value=student.user.first_name)
        ws.cell(row=row, column=2, value=student.user.middle_name)
        ws.cell(row=row, column=3, value=student.user.last_name)
        ws.cell(row=row, column=4, value=student.user.birthday)
        ws.cell(row=row, column=5, value=student.user.email)
        ws.cell(row=row, column=6, value=student.user.phone_number)
        ws.cell(row=row, column=7, value=str(student.school))
        ws.cell(row=row, column=8, value=str(student.grade))
        row += 1
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(f'Студенты {datetime.now().strftime("%d/%m/%y")}.xlsx')
    return response