import secrets
import string

from openpyxl import load_workbook

from .models import School, SchoolContactPersone
from users.models import User
from regions.models import Address, City, TerAdministration

from users.mailing import send_mail

def get_sheet(form):
    workbook = load_workbook(form.cleaned_data['import_file'])
    sheet = workbook.active
    return sheet

def cheak_col_match(sheet, fields_names_set):
    i = 0
    col_count = sheet.max_column
    sheet_fields = []
    sheet_col = {}
    if sheet[f"A2"].value is None:
        return [False, 'EmptySheet']
    try:
        for col_header in range(1, col_count+1):
            if sheet.cell(row=1,column=col_header).value is not None:
                sheet_fields.append(sheet.cell(row=1,column=col_header).value)
                sheet_col[col_header] = sheet.cell(row=1,column=col_header).value
        missing_fields = []
        for field in fields_names_set:
            if field not in sheet_fields:
                missing_fields.append(field)
        if len(missing_fields) != 0:
            return [False, 'FieldError', missing_fields]
    except IndexError:
            return [False, 'IndexError']
    return [True, sheet_col]

def load_worksheet_dict(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            school_name = sheet[f"A{row}"].value
            if school_name != None:
                sheet_dict[fields_names_set[col]].append(sheet.cell(row=row,column=col).value)
    return sheet_dict

def schools(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return [False, 'IndexError']
    
    #Требуемые поля таблицы
    fields_names = {
        'Название школы', 'ИНН школы', 'Теруправление',
        'Город', 'Улица', 'Номер здания',
        'Фамилия', 'Имя', 'Отчество',
        'Email', 'Телефон'
    }

    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names

    sheet_dict = load_worksheet_dict(sheet, cheak_col_names[1])
    schools_count = 0
    problems = []
    dublicates = []
    for row in range(len(sheet_dict['Название школы'])):
        #Загрузка школы
        school = load_school(sheet_dict, row)
        if school[0] == 'Duplicate':
            dublicates.append(school[1])
        elif school[0] != True:
            problems.append(school)
        #Загрузка контакт школы
        if school[0] == True or (school[0] == 'Duplicate' and hasattr(school[1], 'school_contact')):
            school_contact = load_school_contact(school[1], sheet_dict, row)
            if school_contact[0] == True:
                schools_count += 1
            else:
                problems.append(school_contact)
    return [True, schools_count, problems, dublicates]

def load_school(sheet, row):
    try: inn = str(int(sheet["ИНН школы"][row]))
    except: inn = str(sheet["ИНН школы"][row])
    cheak_school = School.objects.filter(inn=inn)
    if len(cheak_school) == 0:
        #Определяем город
        city_name = sheet["Город"][row]
        city_name = city_name.strip()
        city_name = city_name.capitalize()
        city = City.objects.filter(name=city_name)
        if len(city) == 0:
            return['CityErorr', city_name]
        city = city[0]
        
        #Добавляем адрес
        street = sheet["Улица"][row]
        street = street.strip()
        street = street.capitalize()
        building_number = sheet["Номер здания"][row]
        try: building_number = str(int(building_number))
        except: building_number = str(building_number)
        building_number = building_number.replace(" ", "")
        building_number = building_number.upper()
        address = Address.objects.filter(city=city, street=street, building_number=building_number)
        if len(address) == 0:
            address = Address(
                city=city,
                street=street,
                building_number=building_number
            )
            address.save()
        else: address = address[0]
        
        #Определяем теруправление
        admin_name = sheet["Теруправление"][row]
        admin_name = admin_name.strip()
        admin_name = admin_name.capitalize()
        ter_admin = TerAdministration.objects.filter(name=admin_name)
        if len(ter_admin) == 0:
            return['TerAdminErorr', admin_name]
        ter_admin = ter_admin[0]

        #Добавляем школу
        school = School(
            inn=inn,
            name=sheet["Название школы"][row],
            ter_administration=ter_admin,
            address=address,
        )
        school.save()
        return [True, school]
    return ['Duplicate', cheak_school[0]]

def password_generator():
    alphabet = string.ascii_letters + string.digits
    password = ''.join(secrets.choice(alphabet) for i in range(12))
    return password

def load_school_contact(school, sheet, row):
    #Добавляем пользователя
    email = sheet["Email"][row]
    cheak_user = User.objects.filter(email=email)
    if len(cheak_user) == 0:
        email = email.replace(" ", "")
        email = email.lower()
        password = password_generator()
        user = User.objects.create_user(email, password)

        first_name = sheet["Имя"][row]
        first_name = first_name.replace(" ", "")
        first_name = first_name.capitalize()
        user.first_name = first_name
        middle_name = sheet["Отчество"][row]
        middle_name = middle_name.replace(" ", "")
        middle_name = middle_name.capitalize()
        user.middle_name  = middle_name
        last_name = sheet["Фамилия"][row]
        last_name = last_name.replace(" ", "")
        last_name = last_name.capitalize()
        user.last_name = last_name
        try: user.phone_number = str(int(sheet["Телефон"][row]))
        except: user.phone_number = str(int(sheet["Телефон"][row]))
        user.role = 'RSC'
        user.save()

        #Добавляем контакт
        contact = SchoolContactPersone(
            user=user,
            school=school
        )
        contact.save()
        
        #Отправляем email+пароль на почту
        subject = 'Данные для входа в личный кабинет skillsguide.ru'
        html = f'Здравствуйте, {user.first_name}!<p>Вам предоставлен доступ к платформе http://skillsguide.ru/ (проект "Мой выбор"), как представителю школы "{school.name}".</p> <p><br><b>Логин:</b> {user.email}<br><b>Пароль:</b> {password}</p><br><br>Это автоматическое письмо на него не нужно отвечать.'
        text = f'Здравствуйте!\n Здравствуйте, {user.first_name}! \nВам предоставлен доступ к платформе http://skillsguide.ru/ (проект "Мой выбор"), как представителю школы "{school.name}".\nЛогин: {user.email}\nПароль: {password} \n\nЭто автоматическое письмо на него не нужно отвечать.'
        to_name = f"{user.first_name} {user.last_name}"
        to_email = email
        send_mail(subject, html, text, to_name, to_email)

        return [True, contact]
    user = cheak_user[0]
    cheak_contact = SchoolContactPersone.objects.filter(user=user)
    if len(cheak_contact) == 0:
        contact = SchoolContactPersone(
            user=user,
            school=school
        )
        contact.save()
    return ['UserDublicateError', email]