from django.http import HttpResponse
from django.utils.encoding import escape_uri_path

from datetime import datetime
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from .models import TrainingProgram


def programs():
    programs = TrainingProgram.objects.all().prefetch_related(
        'disability_types', 'criteria'
    ).select_related('education_center', 'competence')

    wb = Workbook()
    ws = wb.active
    ws.title = "Професии"
    col_titles = [
        "ЦО", 
        "Программа", 
        "Компетенция",
        "Краткое описание",
        "Ссылка на программу",
        "ОВЗ",
        "Критерий 1",
        "Критерий 2",
        "Критерий 3",
        "Критерий 4",
        "Критерий 5"
    ]
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)

    for row, program in enumerate(programs, start=2):
        disabilities = []
        for disability in program.disability_types.all():
            disabilities.append(disability.name)

        ws.cell(row=row, column=1, value=program.education_center.name)
        ws.cell(row=row, column=2, value=program.name)
        ws.cell(row=row, column=3, value=program.competence.name)
        ws.cell(row=row, column=4, value=program.short_description)
        ws.cell(row=row, column=5, value=program.program_link)
        ws.cell(row=row, column=6, value=', '.join(disabilities))
        for col, criterion in enumerate(program.criteria.all(), start=7):
            ws.cell(row=row, column=col, value=criterion.name)
        
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(
            f'programs_{datetime.now().strftime("%d/%m/%y %H:%M")}.xlsx'
        )
    return response