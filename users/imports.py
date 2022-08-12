from openpyxl import load_workbook
from datetime import datetime, timedelta

from users.models import DisabilityType, User
from schools.models import School, Grade


def get_sheet(form):
    workbook = load_workbook(form.cleaned_data['import_file'])
    sheet = workbook.active
    return sheet

def cheak_col_match(sheet, fields_names_set):
    i = 0
    col_count = sheet.max_column
    sheet_fields = []
    sheet_col = {}
    if sheet[f"A2"].value is None:
        return [False, 'EmptySheet']
    try:
        for col_header in range(1, col_count+1):
            if sheet.cell(row=1,column=col_header).value is not None:
                sheet_fields.append(sheet.cell(row=1,column=col_header).value)
                sheet_col[col_header] = sheet.cell(row=1,column=col_header).value
        missing_fields = []
        for field in fields_names_set:
            if field not in sheet_fields:
                missing_fields.append(field)
        if len(missing_fields) != 0:
            return [False, 'FieldError', missing_fields]
    except IndexError:
            return [False, 'IndexError']
    return [True, sheet_col]

def load_worksheet_dict_slots(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            login = sheet[f"B{row}"].value
            if login != None:
                sheet_dict[fields_names_set[col]].append(sheet.cell(row=row,column=col).value)
    return sheet_dict

def students_import(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return [False, 'IndexError']

    fields_names_set = {
        'Школа', 'Фамилия', 
        'Имя', 'Отчество', 'Дата регистрации', 
        'Email', 'Номер телефона', 'Дата рождения', 
        'Класс', 'Буква класса', 'ОВЗ'
    }

    cheak = cheak_col_match(sheet, fields_names_set)
    if cheak[0] == False:
        return cheak
    
    sheet_dict = load_worksheet_dict_slots(sheet, cheak[1])
    problems = []
    nf_ed_centers = set()
    for row in range(len(sheet_dict['Школа'])):
        test = load_student(sheet_dict, row)
        if test[0] == "school":
            problems.append([row, "нет школы", test[1]])
        elif test[0] == "disability":
             problems.append([row, "ОВЗ не найденно", test[1]])
    return problems

def load_student(sheet_dict, row):
    school = School.objects.filter(name=sheet_dict["Школа"][row])
    if len(school) == 0:
        return ["school", sheet_dict["Школа"][row]]
    school = school[0]
    grade_letter = sheet_dict["Буква класса"][row]
    grade_number = int(sheet_dict["Класс"][row])
    school_class = Grade.objects.filter(school=school, grade=grade_number, grade_letter=grade_letter)
    if len(school_class) != 0:
        school_class = school_class[0]
    else:
        school_class = Grade(
            school=school,
            grade=grade_number,
            grade_letter=grade_letter
        )
        school_class.save()
    date_joined = datetime.strptime(sheet_dict["Дата регистрации"][row],"%d/%m/%Y %X")

    first_name = sheet_dict["Фамилия"][row]
    last_name = sheet_dict["Имя"][row]
    middle_name = sheet_dict["Отчество"][row]
    birthday = datetime.strptime(sheet_dict["Дата рождения"][row],"%Y-%m-%d")
    email = sheet_dict["Email"][row]
    if len(User.objects.filter(email=email)) != 0:
        return ["email", email]
    phone_number = sheet_dict["Номер телефона"][row]
    if len(phone_number) <= 20:
        phone_number = '–'
    
    disability = sheet_dict["ОВЗ"][row]
    if disability != None:
        disability = DisabilityType.objects.filter(name=disability)
        if len(disability) != 0:
            disability = disability[0]
        else:
            return ["disability", sheet_dict["ОВЗ"][row]]
    
    password = 'copp1234'
    user = User.objects.create_user(email, password)
    user.first_name = first_name
    user.middle_name  = middle_name
    user.last_name = last_name
    user.birthday = birthday
    user.phone_number = phone_number
    user.disability_type = disability
    user.code = 'NLOG'
    user.role = 'ST'
    user.relocate_status = 'ST'
    user.school = school
    user.school_class = school_class
    user.date_joined = date_joined
    user.save()
    return ['OK', user]
