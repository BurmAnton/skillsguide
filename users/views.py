import string
import random
import json
from xml.dom import UserDataHandler

from django.shortcuts import render
from django.contrib import auth
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.urls import reverse
from django.db import IntegrityError
#Decorators
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
#Models
from .models import User, DisabilityType, Parent
from schools.models import School, Grade, SchoolContactPersone, SchoolStudent
from schedule.models import Bundle, EducationCenter

from education_centers.forms import ImportDataForm
from .imports import students_import

from .mailing import send_mail
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook


# Create your views here.
@csrf_exempt
def login(request):
    message = None
    if request.user.is_authenticated:
        if len(User.objects.filter(role="SST", email=request.user.email)) != 0:
            return HttpResponseRedirect(reverse('student_profile', args=(request.user.id,)))
        if len(User.objects.filter(role="RSC", email=request.user.email)) != 0:
            contact = SchoolContactPersone.objects.get(user=request.user)
            return HttpResponseRedirect(reverse('school_profile', args=(contact.school.id,)))
        elif len(User.objects.filter(role="TCH", email=request.user.email)) != 0:
            return HttpResponseRedirect(reverse('trainer_profile', args=(request.user.id,1)))
        elif len(User.objects.filter(role="REC", email=request.user.email)) != 0:
            edu_center = EducationCenter.objects.filter(contact_person=request.user)
            if len(edu_center) != 0:
                return HttpResponseRedirect(reverse("ed_center_dashboard", args=(edu_center[0].id,)))
        elif request.user.is_staff:
            return HttpResponseRedirect(reverse("admin:index"))
    elif request.method == "POST":
        email = request.POST["email"]
        password = request.POST["password"]

        user = auth.authenticate(email=email, password=password)
        if user is not None:
            auth.login(request, user)
            return HttpResponseRedirect(reverse("login"))
        else:
            message = "Неверный логин и/или пароль."

    schools = School.objects.all()
    cities = set()
    for school in schools:
        cities.add(school.address.city) 
    return render(request, "user/login.html", {
        "message": message,
        "page_name": "ЦОПП СО | Авторизация",
        'schools': schools,
        'cities': cities,
        'disability_types': DisabilityType.objects.all()
    })

@login_required
def logout(request):
    auth.logout(request)
    return HttpResponseRedirect(reverse("login"))

def code_generator(size=6, chars=string.ascii_uppercase + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))


@login_required
@csrf_exempt
def mailing_form(request):
    message = ""
    if request.method == "POST":
        bundle = Bundle.objects.get(id=request.POST['bundle'])
        schools = School.objects.filter(bundles=bundle)
        users = User.objects.filter(school__in=schools)
        
        for user in users:
            #Отправляем письмо
            text = request.POST['text']
            html = text
            subject = request.POST['subject']
            html = html
            text = text
            to_name = f"{user.first_name} {user.last_name}" 
            to_email = user.email
            send_mail(subject, html, text, to_name, to_email)

        message = "OK"

    return render(request, "user/mailing_form.html", {
        'bundles': Bundle.objects.all(),
        'message': message,
        'users': users
    })


@csrf_exempt
def password_recovery(request, step):
    if request.method == "POST":
        if step == 1:
            data = json.loads(request.body)
            email = data.get("email", "")
            user = User.objects.filter(email=email)
            if len(user) != 0:
                user = user[0]
                code = code_generator()
                user.code = code
                user.save()
                
                #Посылаем письмо с кодом
                subject = 'Востановление пароля skillsguide.ru'
                html = f'Здравствуйте!<p>Вы получили это письмо потому, что Вы (либо кто-то, выдающий себя за вас) попросили выслать новый пароль к вашей учётной записи на сайте http://skillsguide.ru/. <br> Если вы не просили выслать пароль, то не обращайте внимания на это письмо. <br> Код подтверждения для смены пароля: {code} <br> Это автоматическое письмо на него не нужно отвечать.</p>'
                text = f'Здравствуйте!\n Вы получили это письмо потому, что Вы (либо кто-то, выдающий себя за вас) попросили выслать новый пароль к вашей учётной записи на сайте http://skillsguide.ru/. \n Если вы не просили выслать пароль, то не обращайте внимания на это письмо. \n Код подтверждения для смены пароля: {code} \n Это автоматическое письмо на него не нужно отвечать.'
                to_name = f"{user.first_name} {user.last_name}"
                to_email = email
                send_mail(subject, html, text, to_name, to_email)

                return JsonResponse({"message": "Email exist"}, status=201)
            return JsonResponse({"message": "Email not found"}, status=201)
        if step == 2:
            data = json.loads(request.body)
            email = data.get("email", "")
            code = data.get("code", "")
            user = User.objects.get(email=email)
            if user.code == code:
                return JsonResponse({"message": "Code matches"}, status=201)
            return JsonResponse({"message": "Code not matches"}, status=201)
        if step == 3:
            data = json.loads(request.body)
            email = data.get("email", "")
            password = data.get("password", "")
            confirmation = data.get("confirmation", "")
            if password == confirmation:
                user = User.objects.get(email=email)
                user.set_password(password)
                user.save()
                return JsonResponse({"message": "Password changed"}, status=201)
            return JsonResponse({"message": "Passwords mismatch"}, status=201)
    return HttpResponseRedirect(reverse("login"))

@login_required()
@csrf_exempt
def change_password(request):
    if request.method == "POST":
        email = request.user.email
        current_password = request.POST["current_password"]
        password = request.POST["password"]
        confirmation = request.POST["confirmation"]
        user = auth.authenticate(email=email, password=current_password)
        if user is not None:
            if password == confirmation:
                user.set_password(password)
                user.save()
                login(request, user)
    return HttpResponseRedirect(reverse("login"))

@csrf_exempt
def registration(request):
    if request.method == "POST":
        data = json.loads(request.body)
        email = data.get("email", "")
        password = data.get("password", "")
        confirmation = data.get("confirmation", "")

        phone = data.get("phone", "").replace(" - ", "-")
        first_name = data.get("first_name", "")
        last_name = data.get("last_name", "")
        middle_name = data.get("middle_name", "")
        birthday = data.get("birthday", "")

        school_id = data.get("school_id", "")
        grade_number = data.get("grade_number", "")
        grade_letter = data.get("grade_letter", "")
        school = School.objects.get(id=school_id)

        if password != confirmation:
            return JsonResponse({"message": "Password mismatch."}, status=201)
        try:
            user = User.objects.create_user(email, password)
            user.first_name = first_name
            user.middle_name  = middle_name
            user.last_name = last_name
            user.birthday = birthday
            user.phone_number = phone

            user.role = 'SST'
            grade = Grade.objects.filter(
                school=school,
                grade=grade_number,
                grade_letter=grade_letter
            )
            if len(grade) != 0:
                grade = grade[0]
            else:
                grade = Grade(
                    school=school,
                    grade=int(grade_number),
                    grade_letter=grade_letter.upper()
                )
                grade.save()
            user.save()
            disabilities = data.get("disabilities", "")
            user.disability_types.add(*disabilities)
            user.save()
            
            parent_first_name = data.get("parent_first_name", "")
            parent_last_name = data.get("parent_last_name", "")
            parent_middle_name = data.get("parent_middle_name", "")
            if parent_first_name != "":
                parent = Parent(
                    first_name=parent_first_name,
                    last_name=parent_last_name,
                    middle_name=parent_middle_name
                )
                parent.save()
                parent.children.add(user)
                parent.save()
            
            school_student = SchoolStudent(
                user=user,
                school=school,
                grade=grade
            )
            school_student.save()
            
        except IntegrityError:
            return JsonResponse({"message": "Email already taken."}, status=201)

        return JsonResponse({"message": "Account created successfully."}, status=201)
    return HttpResponseRedirect(reverse("login"))

def reg_choice(request):
    return HttpResponseRedirect(reverse("login"))

def reg_stage(request, choice, stage):
    return HttpResponseRedirect(reverse("login"))

@login_required
@csrf_exempt
def import_students_coordinator(request):
    if request.method == "POST":
        form = ImportDataForm(request.POST, request.FILES)
        if form.is_valid():
            message = students_import(form)
        else:
            data = form.errors
        form = ImportDataForm()
        return render(request, "user/import.html",{
            'form': form,
            'message': message
        })

    form = ImportDataForm()
    return render(request, "user/import.html",{
        'form': form
    })


@csrf_exempt
def students_report(request):
    users = User.objects.filter(role='ST')
    wb = Workbook()
    ws = wb.active
    ws.title = "Cтуденты"
    column_names = [
        "Школа","Фамилия", 
        "Имя", "Отчество", 
        "Дата регистрации", "Email", 
        "Номер телефона", "Дата регистрации",
    ]
    active_column = 1
    for name in column_names:
        ws.cell(row=1, column=active_column, value=name)
        active_column += 1
    
    active_row = 2
    for user in users:
        if len(user.bundles.all()) == 0 and len(user.school.bundles.all())!= 0:
                cell_values = {
                    "Школа": user.school.name, 
                    "Фамилия": user.last_name,
                    "Имя": user.first_name, 
                    "Отчество": user.middle_name,
                    "Email": user.email, 
                    "Номер телефона": user.phone_number,
                    "Дата регистрации": str(user.date_joined),
                }
                active_col = 1
                for key, value in cell_values.items():
                    ws.cell(row=active_row, column=active_col, value=value)
                    active_col += 1
                active_row += 1


    wb.template = False
    wb.save('students_list.xlsx')
    response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students_list.xlsx'
    return response